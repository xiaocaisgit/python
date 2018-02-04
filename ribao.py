#coding:utf8
import io
import ftplib
import datetime,time
import sys,os,glob

import openpyxl,pysvn
from openpyxl.styles import PatternFill,Border,Side,Alignment
#vip 就是没有运维的省份
vip = set(['09','27','20'])
pro_dict = {
    '01':'安徽',
    '02':'北京',
    '03':'重庆',
    '04':'福建',
    '05':'甘肃',
    '06':'广东',
    '07':'广西',
    '08':'贵州',
    '09':'海南',
    '10':'河北',
    '11':'河南',
    '12':'黑龙江',
    '13':'湖北',
    '14':'湖南',
    '15':'吉林',
    '16':'江苏',
    '17':'江西',
    '18':'辽宁',
    '19':'内蒙古',
    '20':'宁夏',
    '21':'青海',
    '22':'山东',
    '23':'山西',
    '24':'陕西',
    '25':'上海',
    '26':'四川',
    '27':'天津',
    '28':'新疆',
    '29':'云南',
    '30':'浙江'
}
def excelAry():
    '''
    将日期（一月中的第mday天转换为excel中X轴坐标）
    :param mday
    :return:
    '''
    mday = today.day + 1
    Letter = [chr(x) for x in range(65, 91)]
    x = []
    if mday <= 26:
        x = [Letter[mday % 26 - 1]]
    else:
        x.append(Letter[mday // 26 - 1 ])
        x.append(Letter[mday % 26 - 1 ])
    return ''.join(x)

def GetFtpfiles(host,user,password,rPath,lPath):
    '''
    :param host:
    :param user:
    :param password:
    :param rPath: FTP路径
    :param lPath: 本地路径
    :return: 文件列表
    '''
    try:
        ftp = ftplib.FTP(host,user,password,timeout=3)
    except Exception as e:
        print(e)
        os.system('pause')
        sys.exit(1)

    if ftp.getwelcome():
        print("连接到FTP服务器成功")

    ftp.encoding='gbk'
    try:
        ftp.cwd(rPath)
    except Exception as e:
        if 'No such file or directory' in str(e):
            print('FTP 中没有找到目录: %s,可能未到上传时间' %rPath)
            os.system('pause')
            sys.exit(15)

    print('切换到FTP路径:%s '
          '正在下载日报' %rPath)
    files = ftp.nlst('/%s' %rPath)

    count = len(files)
    for i in range(count):
        with open('/'.join([lPath,files[i]]),'wb') as f:
            ftp.retrbinary("RETR %s" %files[i],f.write)
        print("下载分公司日报中:%s/%s" % ((i + 1),count), end='\r')
    print("日报本地路径:%s" %lPath)
    return files



today = datetime.datetime.now()

yesterday = today - datetime.timedelta(1)
today_stf = today.strftime('%Y-%m-%d')
yesterday_stf = yesterday.strftime('%Y-%m-%d')
try:
    os.makedirs("%s/%s/巡检日报" %(today_stf,today.strftime('%Y%m%d')))
except FileExistsError:
    print("文件或目录已经存在!")
print("开始连接SVN 下载昨天日报...")
svn_server = '192.168.11.253'
url = 'svn://%s/TZManager/BeiJing/R,日报汇总/%s年/%s/%s/' %(svn_server,yesterday.year,yesterday.strftime("%Y-%m"),yesterday_stf)

client = pysvn.Client()
try:
    client.checkout(url,today_stf)
except Exception as e:
    if "doesn't exist" in str(e):
        print(url,"昨天值班同学未提交日报，尝试checkout前一天日报")
        twodays = yesterday-datetime.timedelta(1)
        twodays_stf = twodays.strftime('%Y-%m-%d')
        url = 'svn://%s/TZManager/BeiJing/R,日报汇总/%s年/%s/%s/' %(svn_server,twodays.year,twodays.strftime("%Y-%m"),twodays_stf)
        try:
            client.checkout(url,today_stf)
        except Exception as e:
            print(e)
            os.system('pause')
            sys.exit(12)
print("下载SVN日报文件%s到 %s/" %(url,today_stf))
generate_border = Border(left=Side(style='thin'),
                   right=Side(style='thin'),
                   top=Side(style='thin'),
                   bottom=Side(style='thin'))
generate_alignment = Alignment(horizontal='center',vertical='center')

if today.day == 2:
    dailyreport = '%s\每日分公司故障日报及故障统计%s.xlsx' %(today_stf,today_stf)
    wb = openpyxl.load_workbook(dailyreport,read_only=False)
    ws = wb.active
    ws['A1']= '统计日期'
    for i in pro_dict.keys():
        mycell = ws['%s%s' % ('A', int(i) + 1)]
        mycell.value=pro_dict[i]
        mycell.border = generate_border
        mycell.alignment = generate_alignment
    wb.save(dailyreport)
else:
    dailyreport = glob.glob("%s\每日分公司故障日报及故障统计*.xlsx" % today_stf)
    if len(dailyreport) == 0:
        print("日报文件不存在！")
        os.system('pause')
        sys.exit(13)

    elif len(dailyreport) > 1:
        print("发现多个日报文件", ','.join(dailyreport))
        os.system('pause')
        sys.exit(14)
    else:
        dailyreport = ''.join(dailyreport)
files = GetFtpfiles('10.10.12.12','ribao','ribao','%s\巡检日报' %today.strftime('%Y%m%d'),'%s/%s/巡检日报' %(today_stf,today.strftime('%Y%m%d')))
finished = set(list(map(lambda x:x[0:2],files)))
no_finished = set(pro_dict.keys()).difference(finished).difference(vip)
no_finished.remove('02')
print('未提交日报的省份\n\t',','.join([x for x in list(map(lambda n:pro_dict[n],no_finished))]))
pos_x = excelAry()
wb = openpyxl.load_workbook(dailyreport,read_only=False)
sheet = wb[wb.sheetnames[0]]
sheet['%s1' % (pos_x)] = today_stf
for i in pro_dict.keys():
    mycell = sheet['%s%s' %(pos_x,int(i)+1)]
    mycell.border = generate_border
    mycell.alignment = generate_alignment
    if i in no_finished:
        mycell.value = '未交'
        mycell.fill = PatternFill(fill_type='solid', fgColor='ff0000')
    elif i in vip:
        mycell.value = None
        mycell.fill = PatternFill(fill_type='solid', fgColor='F7F709')
    else:
        mycell.value = '1'
        mycell.fill = PatternFill(fill_type='solid', fgColor='008000')
wb.save('%s\每日分公司故障日报及故障统计%s.xlsx' %(today_stf,today_stf))

print("日报生成成功。请手动清理旧文件，并截取cacti流量图")
os.system('pause')



